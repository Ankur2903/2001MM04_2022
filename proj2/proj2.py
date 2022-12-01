import pandas as pd
import zipfile
import shutil
from zipfile import ZipFile
import os
from pathlib import Path
import streamlit as st
import plotly.express as px
import zipfile
from PIL import Image
import os
from tkinter import N
os.system("cls")
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.borders import Border, Side
from datetime import datetime
start_time = datetime.now()

# set web page title.
st.set_page_config(page_title="Ankur's website")
st.header('Welcome')
first,last = st.columns(2)
# set web page header.

# set web page sub title.
first.subheader('For a single file.....')
file = first.file_uploader("upload a xlsx file",type='xlsx')
mod1 = first.number_input("Enter MOD Value1",0)
last.subheader('For a multiple files.....')
path = last.text_input('Enter path of files')
path.replace('\\','/')
last.header('')
mod2 = last.number_input("Enter MOD Value2",0)
curr  = os.getcwd()#copy current folder adderss in curr
if first.button("Compute1"):
    try:
        wb = load_workbook(file)
        sheet = wb.active
        v = sheet.max_row
        n = np.minimum(v,201)
        time =[]#time list
        u = []#U list
        v = []#V list
        w = []#W list
        u_ = []# u - avg_of_u
        v_ = []#v - avg_of_v
        w_ = []#w - avg_of_w
        octant = []#store octant value
        total = [0]*8#count of 1,2,3,4,-1,-2,-3,-4
        octants = ["Mod Transition Count","","Count","+1","-1","+2","-2","+3","-3","+4","-4"," "," "]
        list_ = [1,-1,2,-2,3,-3,4,-4]
        octants1 = ["+1","-1","+2","-2","3","-3","+4","-4"]
        octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
        for i in range(2, n + 1):
            time.append(sheet.cell(row=i, column=1).value)#append time value in time list
            u.append(sheet.cell(row=i, column=2).value)
            v.append(sheet.cell(row=i, column=3).value)
            w.append(sheet.cell(row=i, column=4).value)

        avg_of_u = np.mean(u)
        avg_of_v = np.mean(v)
        avg_of_w = np.mean(w)#store avg of u,v,w in avg_of_u,avg_of_v,avg_of_w

        for i in range(n-1):
            v_.append(v[i]-avg_of_v)#append V-Vavg in v_
            u_.append(u[i]-avg_of_u)
            w_.append(w[i]-avg_of_w)

        for i in range(n-1):#for identfy octant
            if u_[i]>=0 and v_[i]>=0 and w_[i]>0:
                octant.append(+1)
            elif u_[i]<0 and v_[i]>=0 and w_[i]>0:
                octant.append(+2)
            elif u_[i]<0 and v_[i]<0 and w_[i]>=0:
                octant.append(+3)
            elif u_[i]>=0 and v_[i]<0 and w_[i]>=0:
                octant.append(+4)
            elif u_[i]>=0 and v_[i]>=0 and w_[i]<=0:
                octant.append(-1)
            elif u_[i]<0 and v_[i]>=0 and w_[i]<=0:
                octant.append(-2)
            elif u_[i]<0 and v_[i]<0 and w_[i]<0:
                octant.append(-3)
            else:
                octant.append(-4)

        for j in range(0,n-1):#for count total no of 1,2,3,4,-1,-2,-3,-4
            if octant[j] ==1:
                total[0] = total[0] +1
            elif octant[j] ==2:
                total[2] = total[2] +1
            elif octant[j] ==3:
                total[4] = total[4] +1
            elif octant[j] ==4:
                total[6] = total[6] +1
            elif octant[j] ==-1:
                total[1] = total[1] +1
            elif octant[j] ==-2:
                total[3] = total[3] +1
            elif octant[j] ==-3:
                total[5] = total[5] +1
            else:
                total[7] = total[7] +1
                
        octant_2d = [[0]*((n-2)//mod1+1),[0]*((n-2)//mod1+1),[0]*((n-2)//mod1+1),[0]*((n-2)//mod1+1),[0]*((n-2)//mod1+1),[0]*((n-2)//mod1+1),[0]*((n-2)//mod1+1),[0]*((n-2)//mod1+1)]#store count of 1 in every mod value	
        for i in range(0,(n-2)//mod1):
            for j in range(i*mod1,(i+1)*mod1):
                if octant[j] ==1:#count of 1 in i*mod_1 to (i+1)*mod
                    octant_2d[0][i] = octant_2d[0][i] +1
                elif octant[j] ==2:
                    octant_2d[2][i] = octant_2d[2][i] +1
                elif octant[j] ==3:
                    octant_2d[4][i] = octant_2d[4][i] +1
                elif octant[j] ==4:
                    octant_2d[6][i] = octant_2d[6][i] +1
                elif octant[j] ==-1:
                    octant_2d[1][i] = octant_2d[1][i] +1
                elif octant[j] ==-2:
                    octant_2d[3][i] = octant_2d[3][i] +1
                elif octant[j] ==-3:
                    octant_2d[5][i] = octant_2d[5][i] +1
                else:
                    octant_2d[7][i] = octant_2d[7][i] +1

        for i in range(((n-2)//mod1)*mod1,n-1):
            if octant[i] ==1:#count 1 in rest element 
                octant_2d[0][(n-2)//mod1] = octant_2d[0][(n-2)//mod1] +1
            elif octant[i] ==2:
                octant_2d[2][(n-2)//mod1] = octant_2d[2][(n-2)//mod1] +1
            elif octant[i] ==3:
                octant_2d[4][(n-2)//mod1] = octant_2d[4][(n-2)//mod1] +1
            elif octant[i] ==4:
                octant_2d[6][(n-2)//mod1] = octant_2d[6][(n-2)//mod1] +1
            elif octant[i] ==-1:
                octant_2d[1][(n-2)//mod1] = octant_2d[1][(n-2)//mod1] +1
            elif octant[i] ==-2:
                octant_2d[3][(n-2)//mod1] = octant_2d[3][(n-2)//mod1] +1
            elif octant[i] ==-3:
                octant_2d[5][(n-2)//mod1] = octant_2d[5][(n-2)//mod1] +1
            else:
                octant_2d[7][(n-2)//mod1] = octant_2d[7][(n-2)//mod1] +1

        line1 = []
        for i in range(n-1):#made a list for append -from- in rows list
            if  i%13==0 and i<((n-2)//mod1 +1)*14+2:
                line1.append("From")
            else:
                line1.append("")

        line2 = []
        for i in range(n+2):#made a list for append [1,-1,2,-2,3,-3,4,-4] and range of mod in row list
            if i%13==1 and i<((n-2)//mod1 +2)*13 and i!=1:
                line2.append(str((i//13-1)*mod1)+"-"+str(np.minimum((i//13)*mod1-1,n-2)))#append range
            elif i<((n-2)//mod1 +2)*13:
                line2.append(octants[i%13])#append [1,-1,2,-2,3,-3,4,-4]
            else:
                line2.append("")#append blank

        list_of_max_count = []
        def function1(first, last,a):#define a function for count countinous a,(1,-1,2,-2,3,-3,4,-4)and put them in list named line
            line = [0]*8
            for i in range (first,last):
                if octant[i]==a and octant[i+1]==1:#count countinous a,1
                    line[0] = line[0] +1  
                if octant[i]==a and octant[i+1]==-1:#count countinous a,-1
                    line[1] = line[1] +1
                if octant[i]==a and octant[i+1]==2:#count countinous a,2
                    line[2] = line[2] +1
                if octant[i]==a and octant[i+1]==-2:#count countinous a,-2
                    line[3] = line[3] +1
                if octant[i]==a and octant[i+1]==3:#count countinous a,3
                    line[4] = line[4] +1
                if octant[i]==a and octant[i+1]==-3:#count countinous a,-3
                    line[5] = line[5] +1
                if octant[i]==a and octant[i+1]==4:#count countinous a,4
                    line[6] = line[6] +1
                if octant[i]==a and octant[i+1]==-4:#count countinous a,-4
                    line[7] = line[7] +1
            list_of_max_count.append(max(line))
            return line

        linenext = [[],[],[],[],[],[],[],[]]
        linenext[0].append("To")#storing all list in 2dlist for overall mod transition count
        for k in range(1,8):
            linenext[k].append("")
        for k in range(8):
            linenext[k].append(octants1[k])
        for j in range(8):
            list1 =  function1(0,n-2,int(octants1[j]))
            for z in range(8):
                linenext[z].append(list1[z])
        for k in range(8):
            linenext[k].append("")
            linenext[k].append("")
            linenext[k].append("")

        for i in range((n-2)//mod1+1):#storing all list in 2dlist for overall mod transition count
            linenext[0].append("To")
            for k in range(1,8):
                linenext[k].append("")
            for k in range(8):
                linenext[k].append(octants1[k])
            for j in range(8):
                list1 =  function1(i*mod1,np.minimum((i+1)*mod1,n-2),int(octants1[j]))
                for z in range(8):
                    linenext[z].append(list1[z])
            for k in range(8):
                linenext[k].append("")
                linenext[k].append("")
                linenext[k].append("")

        for i in range(n):
            for k in range(8):
                linenext[k].append("")

        max_count = [0]*8#list of Longest Subsquence Length element all 0
        counts = [0]*8#list of count of Longest Subsquence Length
        for i in range(8):
            count = 0#count start from zero
            for j in range(n-2):
                if octant[j]==int(octants1[i]):
                    count = count+1#counting every element
                else:
                    if count>max_count[i]:#for greater count update max count and counts
                        max_count[i] = count
                        counts[i] = 1
                    elif count==max_count[i]:#for eqal count update counts
                        counts[i]=counts[i]+1
                        count = 0#for not eqal to privious element make count zero
							
        rank_1 = []
        rank_total = []#ranking for total count
        for i in range(8):
            a =1
            for j in range(8):
                if total[i]<total[j]:
                    a = a+1
            if a==1:
                rank_1.append(list_[i])
            rank_total.append(a)

        max_count = [0]*8#list of Longest Subsquence Length element all 0
        counts = [0]*8#list of count of Longest Subsquence Length
        l1 = []#list for count part of tut04
        l2 = []#list for longest subsequebce part and start time of tut04
        l3 = []#list for count part and end time of tut04

        for i in range(8):
            li =[]#list for storing start and end time for count
            count = 0#count start from zero
            for j in range(n-2):
                if octant[j]==int(octants1[i]):
                    count = count+1#counting every element
                else:
                    if count>max_count[i]:#for greater count update max count and counts
                        max_count[i] = count
                        counts[i] = 1
                        li =[]#blank list becouse max count is changed
                        li.append(time[j-count])#appendint start time
                        li.append(time[j-1])#appendint end time
                    elif count==max_count[i]:#for eqal count update counts
                        counts[i]=counts[i]+1
                        li.append(time[j-count])#appendint start time
                        li.append(time[j-1])#appendint end time
                    count = 0#for not eqal to privious element make count zero
            l1.append(octants1[i])#appending octants in l1
            l1.append("Time")#appending string Time in l1
            l2.append(max_count[i]) #appending max_count in l1
            l2.append("From") #appending string From in l1
            l3.append(counts[i])#appending counts in l1
            l3.append("To")#appending string To in l1
            for k in range(0,counts[i],1):
                l1.append("")
                l2.append(li[2*k])#appendint start time in l2
                l3.append(li[2*k+1])#appendint end time in l3

        f = len(l1)			
        count_ = [0]*8# list for count of Rank 1 Mod Value
        rank_2d = []#2d list for ranking in every mod gap
        for i in range(((n-2)//mod1)+1):#for every mod interval
            rank_mod = []
            for j in range(8):#take a value for compare
                a = 1
                for k in range(8):#comparing with other value
                    if octant_2d[j][i]<octant_2d[k][i]:
                        a = a+1#increasing rank
                if a==1:
                    rank_1.append(list_[j])#storing rank 1 octant
                    count_[j] = count_[j] +1#counting rank 1 octant
                rank_mod.append(a)
            rank_2d.append(rank_mod)

        for i in range(n):
            l1.append("")
            l2.append("")
            l3.append("")
            max_count.append("")
            counts.append("")
            octants1.append("")

        rows = [["","","","","","","","","","","","","","Overall Octant Count","","","","","","","","","","","","","","","","","","","","","Overall Transition Count","","","","","","","","","","Longest Subsquence Length","","","","Longest Subsquence Length with Range",],["Time","U","V","W","U Avg","V Avg","W Avg","U'=U - U avg","V'=V - V avg","W'=W - w avg","Ocatant","","","","","","","","","","","","","","","","","","","","","","","","","To"],[time[0],u[0],v[0],w[0],avg_of_u,avg_of_v,avg_of_w,u_[0],v_[0],w_[0],octant[0],"","","Octant ID","+1","-1","+2","-2","+3","-3","+4","-4","Rank Octant 1","Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank 1 Octant ID","Rank 1 Octant Name","","","octant","+1","-1","+2","-2","+3","-3","+4","-4","","Octant","Longest Subsquence Length","Count","","Octant","Longest Subsquence Length","Count"]]#made 2d list
        from openpyxl import Workbook
        book = Workbook()
        sheet = book.active

        for i in range(n-2):
            if i==0:#append 2nd line in rows
                rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i+1],v_[i+1],w_[i+1],octant[i+1],"","Mod"+str(mod1),"Overall Count",total[0],total[1],total[2],total[3],total[4],total[5],total[6],total[7],rank_total[0],rank_total[1],rank_total[2],rank_total[3],rank_total[4],rank_total[5],rank_total[6],rank_total[7],rank_1[0],octant_name_id_mapping[str(rank_1[0])],"",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
            elif i>=1 and i<=1+(n-2)//mod1:
                rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i],v_[i],w_[i],octant[i],"","",str((i-1)*mod1)+"-"+str(np.minimum((i)*mod1-1,n-2)),octant_2d[0][i-1],octant_2d[1][i-1],octant_2d[2][i-1],octant_2d[3][i-1],octant_2d[4][i-1],octant_2d[5][i-1],octant_2d[6][i-1],octant_2d[7][i-1],rank_2d[i-1][0],rank_2d[i-1][1],rank_2d[i-1][2],rank_2d[i-1][3],rank_2d[i-1][4],rank_2d[i-1][5],rank_2d[i-1][6],rank_2d[i-1][7],rank_1[i],octant_name_id_mapping[str(rank_1[i])],"",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
            elif i==3+(n-2)//mod1:
                rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i],v_[i],w_[i],octant[i],"","","","","","","","","","","","","","","","","","Octant ID","Octant Name","Count of Rank 1 Mod Values","","",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
            elif i>=4+(n-2)//mod1 and i<=11+(n-2)//mod1:
                rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i],v_[i],w_[i],octant[i],"","","","","","","","","","","","","","","","","",list_[i-(4+(n-2)//mod1)],octant_name_id_mapping[str(list_[i-(4+(n-2)//mod1)])],count_[i-(4+(n-2)//mod1)],"","",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
            else:
                rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i],v_[i],w_[i],octant[i],"","","","","","","","","","","","","","","","","","","","","","",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
    
        for row in rows:
            sheet.append(row)
        fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        sheet.conditional_formatting.add(f"W4:AE{2+(n-2)//mod1+4}", CellIsRule(operator='equal', formula=[1], fill=fill))
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
        p=4
        for i in range(2+(n-2)//mod1):
            for k in range(35,44):
                sheet.cell(row=p-1, column=k).border = thin_border
            for j in range(8):
                for k in range(35,44):
                    sheet.cell(row=p, column=k).border = thin_border
                sheet.conditional_formatting.add(f"AJ{p}:AQ{p}", CellIsRule(operator='equal', formula=[list_of_max_count[i*8+j]], fill=fill))
                p=p+1
            p=p+5
        for i in range(3,6+(n-2)//mod1):
            for j in range(14,33):
                sheet.cell(row=i, column=j).border = thin_border
        for i in range(7+(n-2)//mod1,16+(n-2)//mod1):
            for j in range(29,32):
                sheet.cell(row=i, column=j).border = thin_border
        for i in range(3,12):
            for j in range(45,48):
                sheet.cell(row=i, column=j).border = thin_border	
        for i in range(3,f+4):
            for j in range(49,52):
                sheet.cell(row=i, column=j).border = thin_border
        curr  = os.getcwd()#copy current folder adderss in curr
        if os.path.exists("output"):
            for f in os.listdir("output"):
                os.remove(os.path.join("output",f))#delete all file in output folder
            os.rmdir("output")#delete output folder
        os.mkdir(curr.replace('\\','/')+"/output/")
        os.chdir("output")
        book.save("output_file.xlsx")
        with zipfile.ZipFile('files.zip','w') as my_zip:
            my_zip.write('output_file.xlsx')
        os.chdir(curr)
        if os.path.exists("files"):
            os.remove('files')
        p1 = curr + '\\output\\files.zip'
        p2 = curr + '\\files.zip'
        shutil.copy(p1, p2)
        if os.path.exists("output"):
            for f in os.listdir("output"):
                os.remove(os.path.join("output",f))#delete all file in output folder
            os.rmdir("output")#delete output folder
        first,last = st.columns(2)
        with open('files.zip','rb') as fp:
            btn = first.download_button(
                label='download zip file',
                data = fp,
                file_name = 'files.zip',
                mime = 'application/zip'
            )
    except:
        print("part one does not work")
    

if last.button("Compute2"):
    try:
        if os.path.exists("output"):
            for f in os.listdir("output"):
                os.remove(os.path.join("output",f))#delete all file in output folder
            os.rmdir("output")#delete output folder
        os.mkdir(curr.replace('\\','/')+"/output/")
        os.chdir(path)
        file_li = os. listdir()#storing all file name of input folder in list
        print(file_li)
        os.chdir(curr)
        for list in file_li:
            if list[-4:]!='xlsx':
                continue
            os.chdir(path)
            wb = load_workbook(list)
            sheet = wb.active
            print("working on " + list +"...")
            v = sheet.max_row
            n = np.minimum(v,201)
            time =[]#time list
            u = []#U list
            v = []#V list
            w = []#W list
            u_ = []# u - avg_of_u
            v_ = []#v - avg_of_v
            w_ = []#w - avg_of_w
            octant = []#store octant value
            total = [0]*8#count of 1,2,3,4,-1,-2,-3,-4
            octants = ["Mod Transition Count","","Count","+1","-1","+2","-2","+3","-3","+4","-4"," "," "]
            list_ = [1,-1,2,-2,3,-3,4,-4]
            octants1 = ["+1","-1","+2","-2","3","-3","+4","-4"]
            octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
            
            for i in range(2, n + 1):
                time.append(sheet.cell(row=i, column=1).value)#append time value in time list
                u.append(sheet.cell(row=i, column=2).value)
                v.append(sheet.cell(row=i, column=3).value)
                w.append(sheet.cell(row=i, column=4).value)

            avg_of_u = np.mean(u)
            avg_of_v = np.mean(v)
            avg_of_w = np.mean(w)#store avg of u,v,w in avg_of_u,avg_of_v,avg_of_w
            
            for i in range(n-1):
                v_.append(v[i]-avg_of_v)#append V-Vavg in v_
                u_.append(u[i]-avg_of_u)
                w_.append(w[i]-avg_of_w)

            for i in range(n-1):#for identfy octant
                if u_[i]>=0 and v_[i]>=0 and w_[i]>0:
                    octant.append(+1)
                elif u_[i]<0 and v_[i]>=0 and w_[i]>0:
                    octant.append(+2)
                elif u_[i]<0 and v_[i]<0 and w_[i]>=0:
                    octant.append(+3)
                elif u_[i]>=0 and v_[i]<0 and w_[i]>=0:
                    octant.append(+4)
                elif u_[i]>=0 and v_[i]>=0 and w_[i]<=0:
                    octant.append(-1)
                elif u_[i]<0 and v_[i]>=0 and w_[i]<=0:
                    octant.append(-2)
                elif u_[i]<0 and v_[i]<0 and w_[i]<0:
                    octant.append(-3)
                else:
                    octant.append(-4)

            for j in range(0,n-1):#for count total no of 1,2,3,4,-1,-2,-3,-4
                if octant[j] ==1:
                    total[0] = total[0] +1
                elif octant[j] ==2:
                    total[2] = total[2] +1
                elif octant[j] ==3:
                    total[4] = total[4] +1
                elif octant[j] ==4:
                    total[6] = total[6] +1
                elif octant[j] ==-1:
                    total[1] = total[1] +1
                elif octant[j] ==-2:
                    total[3] = total[3] +1
                elif octant[j] ==-3:
                    total[5] = total[5] +1
                else:
                    total[7] = total[7] +1

            octant_2d = [[0]*((n-2)//mod2+1),[0]*((n-2)//mod2+1),[0]*((n-2)//mod2+1),[0]*((n-2)//mod2+1),[0]*((n-2)//mod2+1),[0]*((n-2)//mod2+1),[0]*((n-2)//mod2+1),[0]*((n-2)//mod2+1)]#store count of 1 in every mod value	
            for i in range(0,(n-2)//mod2):
                for j in range(i*mod2,(i+1)*mod2):
                    if octant[j] ==1:#count of 1 in i*mod_1 to (i+1)*mod
                        octant_2d[0][i] = octant_2d[0][i] +1
                    elif octant[j] ==2:
                        octant_2d[2][i] = octant_2d[2][i] +1
                    elif octant[j] ==3:
                        octant_2d[4][i] = octant_2d[4][i] +1
                    elif octant[j] ==4:
                        octant_2d[6][i] = octant_2d[6][i] +1
                    elif octant[j] ==-1:
                        octant_2d[1][i] = octant_2d[1][i] +1
                    elif octant[j] ==-2:
                        octant_2d[3][i] = octant_2d[3][i] +1
                    elif octant[j] ==-3:
                        octant_2d[5][i] = octant_2d[5][i] +1
                    else:
                        octant_2d[7][i] = octant_2d[7][i] +1

            for i in range(((n-2)//mod2)*mod2,n-1):
                if octant[i] ==1:#count 1 in rest element 
                    octant_2d[0][(n-2)//mod2] = octant_2d[0][(n-2)//mod2] +1
                elif octant[i] ==2:
                    octant_2d[2][(n-2)//mod2] = octant_2d[2][(n-2)//mod2] +1
                elif octant[i] ==3:
                    octant_2d[4][(n-2)//mod2] = octant_2d[4][(n-2)//mod2] +1
                elif octant[i] ==4:
                    octant_2d[6][(n-2)//mod2] = octant_2d[6][(n-2)//mod2] +1
                elif octant[i] ==-1:
                    octant_2d[1][(n-2)//mod2] = octant_2d[1][(n-2)//mod2] +1
                elif octant[i] ==-2:
                    octant_2d[3][(n-2)//mod2] = octant_2d[3][(n-2)//mod2] +1
                elif octant[i] ==-3:
                    octant_2d[5][(n-2)//mod2] = octant_2d[5][(n-2)//mod2] +1
                else:
                    octant_2d[7][(n-2)//mod2] = octant_2d[7][(n-2)//mod2] +1

            line1 = []
            for i in range(n-1):#made a list for append -from- in rows list
                if  i%13==0 and i<((n-2)//mod2 +1)*14+2:
                    line1.append("From")
                else:
                    line1.append("")    
            
            line2 = []
            for i in range(n+2):#made a list for append [1,-1,2,-2,3,-3,4,-4] and range of mod in row list
                if i%13==1 and i<((n-2)//mod2 +2)*13 and i!=1:
                    line2.append(str((i//13-1)*mod2)+"-"+str(np.minimum((i//13)*mod2-1,n-2)))#append range
                elif i<((n-2)//mod2 +2)*13:
                    line2.append(octants[i%13])#append [1,-1,2,-2,3,-3,4,-4]
                else:
                    line2.append("")#append blank
        
            list_of_max_count = []
            def function1(first, last,a):#define a function for count countinous a,(1,-1,2,-2,3,-3,4,-4)and put them in list named line
                line = [0]*8
                for i in range (first,last):
                    if octant[i]==a and octant[i+1]==1:#count countinous a,1
                        line[0] = line[0] +1  
                    if octant[i]==a and octant[i+1]==-1:#count countinous a,-1
                        line[1] = line[1] +1
                    if octant[i]==a and octant[i+1]==2:#count countinous a,2
                        line[2] = line[2] +1
                    if octant[i]==a and octant[i+1]==-2:#count countinous a,-2
                        line[3] = line[3] +1
                    if octant[i]==a and octant[i+1]==3:#count countinous a,3
                        line[4] = line[4] +1
                    if octant[i]==a and octant[i+1]==-3:#count countinous a,-3
                        line[5] = line[5] +1
                    if octant[i]==a and octant[i+1]==4:#count countinous a,4
                        line[6] = line[6] +1
                    if octant[i]==a and octant[i+1]==-4:#count countinous a,-4
                        line[7] = line[7] +1
                list_of_max_count.append(max(line))
                return line

            linenext = [[],[],[],[],[],[],[],[]]
            linenext[0].append("To")#storing all list in 2dlist for overall mod transition count
            for k in range(1,8):
                linenext[k].append("")
            for k in range(8):
                linenext[k].append(octants1[k])
            for j in range(8):
                list1 =  function1(0,n-2,int(octants1[j]))
                for z in range(8):
                    linenext[z].append(list1[z])
            for k in range(8):
                linenext[k].append("")
                linenext[k].append("")
                linenext[k].append("")

            for i in range((n-2)//mod2+1):#storing all list in 2dlist for overall mod transition count
                linenext[0].append("To")
                for k in range(1,8):
                    linenext[k].append("")
                for k in range(8):
                    linenext[k].append(octants1[k])
                for j in range(8):
                    list1 =  function1(i*mod2,np.minimum((i+1)*mod2,n-2),int(octants1[j]))
                    for z in range(8):
                        linenext[z].append(list1[z])
                for k in range(8):
                    linenext[k].append("")
                    linenext[k].append("")
                    linenext[k].append("")

            for i in range(n):
                for k in range(8):
                    linenext[k].append("")

            max_count = [0]*8#list of Longest Subsquence Length element all 0
            counts = [0]*8#list of count of Longest Subsquence Length
            for i in range(8):
                count = 0#count start from zero
                for j in range(n-2):
                    if octant[j]==int(octants1[i]):
                        count = count+1#counting every element
                    else:
                        if count>max_count[i]:#for greater count update max count and counts
                            max_count[i] = count
                            counts[i] = 1
                        elif count==max_count[i]:#for eqal count update counts
                            counts[i]=counts[i]+1
                            count = 0#for not eqal to privious element make count zero
                        
            rank_1 = []
            rank_total = []#ranking for total count
            for i in range(8):
                a =1
                for j in range(8):
                    if total[i]<total[j]:
                        a = a+1
                if a==1:
                    rank_1.append(list_[i])
                rank_total.append(a)

            max_count = [0]*8#list of Longest Subsquence Length element all 0
            counts = [0]*8#list of count of Longest Subsquence Length
            l1 = []#list for count part of tut04
            l2 = []#list for longest subsequebce part and start time of tut04
            l3 = []#list for count part and end time of tut04
            for i in range(8):
                li =[]#list for storing start and end time for count
                count = 0#count start from zero
                for j in range(n-2):
                    if octant[j]==int(octants1[i]):
                        count = count+1#counting every element
                    else:
                        if count>max_count[i]:#for greater count update max count and counts
                            max_count[i] = count
                            counts[i] = 1
                            li =[]#blank list becouse max count is changed
                            li.append(time[j-count])#appendint start time
                            li.append(time[j-1])#appendint end time
                        elif count==max_count[i]:#for eqal count update counts
                            counts[i]=counts[i]+1
                            li.append(time[j-count])#appendint start time
                            li.append(time[j-1])#appendint end time
                        count = 0#for not eqal to privious element make count zero
                l1.append(octants1[i])#appending octants in l1
                l1.append("Time")#appending string Time in l1
                l2.append(max_count[i]) #appending max_count in l1
                l2.append("From") #appending string From in l1
                l3.append(counts[i])#appending counts in l1
                l3.append("To")#appending string To in l1
                for k in range(0,counts[i],1):
                    l1.append("")
                    l2.append(li[2*k])#appendint start time in l2
                    l3.append(li[2*k+1])#appendint end time in l3

            f = len(l1)			
            count_ = [0]*8# list for count of Rank 1 Mod Value
            rank_2d = []#2d list for ranking in every mod gap
            for i in range(((n-2)//mod2)+1):#for every mod interval
                rank_mod = []
                for j in range(8):#take a value for compare
                    a = 1
                    for k in range(8):#comparing with other value
                        if octant_2d[j][i]<octant_2d[k][i]:
                            a = a+1#increasing rank
                    if a==1:
                        rank_1.append(list_[j])#storing rank 1 octant
                        count_[j] = count_[j] +1#counting rank 1 octant
                    rank_mod.append(a)
                rank_2d.append(rank_mod)

            for i in range(n):
                l1.append("")
                l2.append("")
                l3.append("")
                max_count.append("")
                counts.append("")
                octants1.append("")
            

            rows = [["","","","","","","","","","","","","","Overall Octant Count","","","","","","","","","","","","","","","","","","","","","Overall Transition Count","","","","","","","","","","Longest Subsquence Length","","","","Longest Subsquence Length with Range",],["Time","U","V","W","U Avg","V Avg","W Avg","U'=U - U avg","V'=V - V avg","W'=W - w avg","Ocatant","","","","","","","","","","","","","","","","","","","","","","","","","To"],[time[0],u[0],v[0],w[0],avg_of_u,avg_of_v,avg_of_w,u_[0],v_[0],w_[0],octant[0],"","","Octant ID","+1","-1","+2","-2","+3","-3","+4","-4","Rank Octant 1","Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank 1 Octant ID","Rank 1 Octant Name","","","octant","+1","-1","+2","-2","+3","-3","+4","-4","","Octant","Longest Subsquence Length","Count","","Octant","Longest Subsquence Length","Count"]]#made 2d list
            os.chdir(curr)
            os.chdir("output")
            from openpyxl import Workbook
            book = Workbook()
            sheet = book.active

            for i in range(n-2):
                if i==0:#append 2nd line in rows
                    rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i+1],v_[i+1],w_[i+1],octant[i+1],"","Mod"+str(mod2),"Overall Count",total[0],total[1],total[2],total[3],total[4],total[5],total[6],total[7],rank_total[0],rank_total[1],rank_total[2],rank_total[3],rank_total[4],rank_total[5],rank_total[6],rank_total[7],rank_1[0],octant_name_id_mapping[str(rank_1[0])],"",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
                elif i>=1 and i<=1+(n-2)//mod2:
                    rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i],v_[i],w_[i],octant[i],"","",str((i-1)*mod2)+"-"+str(np.minimum((i)*mod2-1,n-2)),octant_2d[0][i-1],octant_2d[1][i-1],octant_2d[2][i-1],octant_2d[3][i-1],octant_2d[4][i-1],octant_2d[5][i-1],octant_2d[6][i-1],octant_2d[7][i-1],rank_2d[i-1][0],rank_2d[i-1][1],rank_2d[i-1][2],rank_2d[i-1][3],rank_2d[i-1][4],rank_2d[i-1][5],rank_2d[i-1][6],rank_2d[i-1][7],rank_1[i],octant_name_id_mapping[str(rank_1[i])],"",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
                elif i==3+(n-2)//mod2:
                    rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i],v_[i],w_[i],octant[i],"","","","","","","","","","","","","","","","","","Octant ID","Octant Name","Count of Rank 1 Mod Values","","",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
                elif i>=4+(n-2)//mod2 and i<=11+(n-2)//mod2:
                    rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i],v_[i],w_[i],octant[i],"","","","","","","","","","","","","","","","","",list_[i-(4+(n-2)//mod2)],octant_name_id_mapping[str(list_[i-(4+(n-2)//mod2)])],count_[i-(4+(n-2)//mod2)],"","",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
                else:
                    rows.append([time[i+1],u[i+1],v[i+1],w[i+1],"","","",u_[i],v_[i],w_[i],octant[i],"","","","","","","","","","","","","","","","","","","","","","",line1[i],line2[i+3],linenext[0][i+2],linenext[1][i+2],linenext[2][i+2],linenext[3][i+2],linenext[4][i+2],linenext[5][i+2],linenext[6][i+2],linenext[7][i+2],"",octants1[i],max_count[i],counts[i],"",l1[i],l2[i],l3[i]])
            for row in rows:
                sheet.append(row)
            fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
            sheet.conditional_formatting.add(f"W4:AE{2+(n-2)//mod2+4}", CellIsRule(operator='equal', formula=[1], fill=fill))
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
            p=4
            for i in range(2+(n-2)//mod2):
                for k in range(35,44):
                    sheet.cell(row=p-1, column=k).border = thin_border
                for j in range(8):
                    for k in range(35,44):
                        sheet.cell(row=p, column=k).border = thin_border
                    sheet.conditional_formatting.add(f"AJ{p}:AQ{p}", CellIsRule(operator='equal', formula=[list_of_max_count[i*8+j]], fill=fill))
                    p=p+1
                p=p+5
            for i in range(3,6+(n-2)//mod2):
                for j in range(14,33):
                    sheet.cell(row=i, column=j).border = thin_border
            for i in range(7+(n-2)//mod2,16+(n-2)//mod2):
                for j in range(29,32):
                    sheet.cell(row=i, column=j).border = thin_border
            for i in range(3,12):
                for j in range(45,48):
                    sheet.cell(row=i, column=j).border = thin_border	
            for i in range(3,f+4):
                for j in range(49,52):
                    sheet.cell(row=i, column=j).border = thin_border
            book.save(list[:-5]+" cm_vel_octant_analysis_mod_"+str(mod2)+".xlsx")
            with zipfile.ZipFile('files.zip','a',compression=zipfile.ZIP_DEFLATED) as my_zip:
                my_zip.write(list[:-5]+" cm_vel_octant_analysis_mod_"+str(mod2)+".xlsx")
            os.chdir(curr)
            print("done.")
        if os.path.exists("files"):
            os.remove('files')
        p1 = curr + '\\output\\files.zip'
        p2 = curr + '\\files.zip'
        shutil.copy(p1, p2)
        if os.path.exists("output"):
            for f in os.listdir("output"):
                os.remove(os.path.join("output",f))#delete all file in output folder
            os.rmdir("output")#delete output folder
        first,last = st.columns(2)
        with open('files.zip','rb') as fp:
            btn = last.download_button(
                label='download zip file',
                data = fp,
                file_name = 'files.zip',
                mime = 'application/zip'
            )
    except:
        print('part 2 does not work')


#st.subheader()