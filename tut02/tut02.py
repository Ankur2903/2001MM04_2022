import numpy as np
from openpyxl import load_workbook
def octact_transition_count(mod=5000):
    try:
        wb = load_workbook("input_octant_transition_identify.xlsx")
        sheet = wb["Sheet1"]
        n = sheet.max_row
        #n = row_count
        time =[]#time list
        u = []#U list
        v = []#V list
        w = []#W list
        u_ = []# u - avg_of_u
        v_ = []#v - avg_of_v
        w_ = []#w - avg_of_w
        octant = []#store octant value
        total = [0]*8#count of 1,2,3,4,-1,-2,-3,-4
        octants = ["Mod Transition Count","","Count","+1","-1","+2","-2","+3","-3","+4","-4"," "," "]
        list_ = [1,-1,2,-2,3,-3,4,-4]

        for i in range(2, n + 1):
            time.append(sheet.cell(row=i, column=1).value)#append time value in time list
            u.append(sheet.cell(row=i, column=2).value)
            v.append(sheet.cell(row=i, column=3).value)
            w.append(sheet.cell(row=i, column=4).value)

        avg_of_u = np.mean(u)
        avg_of_v = np.mean(v)
        avg_of_w = np.mean(w)#store avg of u,v,w in avg_of_u,avg_of_v,avg_of_w

        for i in range(n-1):
            v_.append(v[i]-avg_of_v)#append V-Vavg in v_
            u_.append(u[i]-avg_of_u)
            w_.append(w[i]-avg_of_w)
        
        for i in range(n-1):#for identfy octant
                if u_[i]>=0 and v_[i]>=0 and w_[i]>0:
                    octant.append(+1)
                elif u_[i]<0 and v_[i]>=0 and w_[i]>0:
                    octant.append(+2)
                elif u_[i]<0 and v_[i]<0 and w_[i]>=0:
                    octant.append(+3)
                elif u_[i]>=0 and v_[i]<0 and w_[i]>=0:
                    octant.append(+4)
                elif u_[i]>=0 and v_[i]>=0 and w_[i]<=0:
                    octant.append(-1)
                elif u_[i]<0 and v_[i]>=0 and w_[i]<=0:
                    octant.append(-2)
                elif u_[i]<0 and v_[i]<0 and w_[i]<0:
                    octant.append(-3)
                else:
                    octant.append(-4)

        for j in range(0,n-1):#for count total no of 1,2,3,4,-1,-2,-3,-4
            if octant[j] ==1:
                total[0] = total[0] +1
            elif octant[j] ==2:
                total[2] = total[2] +1
            elif octant[j] ==3:
                total[4] = total[4] +1
            elif octant[j] ==4:
                total[6] = total[6] +1
            elif octant[j] ==-1:
                total[1] = total[1] +1
            elif octant[j] ==-2:
                total[3] = total[3] +1
            elif octant[j] ==-3:
                total[5] = total[5] +1
            else:
                total[7] = total[7] +1

        octant_1 = [0]*((n-2)//mod+1)#store count of 1 in every mod value
        octant_2 = [0]*((n-2)//mod+1)
        octant_3 = [0]*((n-2)//mod+1)
        octant_4 = [0]*((n-2)//mod+1)
        octant__1 = [0]*((n-2)//mod+1)
        octant__2 = [0]*((n-2)//mod+1)
        octant__3 = [0]*((n-2)//mod+1)
        octant__4 = [0]*((n-2)//mod+1)
            
        for i in range(0,(n-2)//mod):
            for j in range(i*mod,(i+1)*mod):
                if octant[j] ==1:#count of 1 in i*mod_1 to (i+1)*mod
                    octant_1[i] = octant_1[i] +1
                elif octant[j] ==2:
                    octant_2[i] = octant_2[i] +1
                elif octant[j] ==3:
                    octant_3[i] = octant_3[i] +1
                elif octant[j] ==4:
                    octant_4[i] = octant_4[i] +1
                elif octant[j] ==-1:
                    octant__1[i] = octant__1[i] +1
                elif octant[j] ==-2:
                    octant__2[i] = octant__2[i] +1
                elif octant[j] ==-3:
                    octant__3[i] = octant__3[i] +1
                else:
                    octant__4[i] = octant__4[i] +1

        for i in range(((n-2)//mod)*mod,n-1):
            if octant[i] ==1:#count 1 in rest element 
                octant_1[(n-2)//mod] = octant_1[(n-2)//mod] +1
            elif octant[i] ==2:
                octant_2[(n-2)//mod] = octant_2[(n-2)//mod] +1
            elif octant[i] ==3:
                octant_3[(n-2)//mod] = octant_3[(n-2)//mod] +1
            elif octant[i] ==4:
                octant_4[(n-2)//mod] = octant_4[(n-2)//mod] +1
            elif octant[i] ==-1:
                octant__1[(n-2)//mod] = octant__1[(n-2)//mod] +1
            elif octant[i] ==-2:
                octant__2[(n-2)//mod] = octant__2[(n-2)//mod] +1
            elif octant[i] ==-3:
                octant__3[(n-2)//mod] = octant__3[(n-2)//mod] +1
            else:
                octant__4[(n-2)//mod] = octant__4[(n-2)//mod] +1

        line1 = []
        for i in range(n-1):#made a list for append -from- in rows list
            if  i%13==0 and i<((n-2)//mod +2)*14+2:
                line1.append("From")
            else:
                line1.append("")    

        line2 = []
        for i in range(n-1):#made a list for append [1,-1,2,-2,3,-3,4,-4] and range of mod in row list
            if i%13==1 and i<((n-2)//mod +2)*13 and i!=1:
                line2.append(str((i//13-1)*mod)+"-"+str(np.minimum((i//13)*mod-1,n-2)))#append range
            elif i<((n-2)//mod +2)*13:
                line2.append(octants[i%13])#append [1,-1,2,-2,3,-3,4,-4]
            else:
                line2.append("")#append blank

        def function1(first, last,a):#define a function for count countinous a,(1,-1,2,-2,3,-3,4,-4)and put them in list named line
            line = [0]*8
            for i in range (first,last):
                if octant[i]==a and octant[i+1]==1:#count countinous a,1
                    line[0] = line[0] +1  
                if octant[i]==a and octant[i+1]==-1:#count countinous a,-1
                    line[1] = line[1] +1
                if octant[i]==a and octant[i+1]==2:#count countinous a,2
                    line[2] = line[2] +1
                if octant[i]==a and octant[i+1]==-2:#count countinous a,-2
                    line[3] = line[3] +1
                if octant[i]==a and octant[i+1]==3:#count countinous a,3
                    line[4] = line[4] +1
                if octant[i]==a and octant[i+1]==-3:#count countinous a,-3
                    line[5] = line[5] +1
                if octant[i]==a and octant[i+1]==4:#count countinous a,4
                    line[6] = line[6] +1
                if octant[i]==a and octant[i+1]==-4:#count countinous a,-4
                    line[7] = line[7] +1
            return line

        from openpyxl import Workbook
        book = Workbook()
        sheet = book.active

        rows = [
            ["Time","U","V","W","U Avg","V Avg","W Avg","U'=U - U avg","V'=V - V avg","W'=W - w avg","Ocatant","","","+1","-1","+2","-2","+3","-3","+4","-4"]
        ]#made 2d list

        for i in range(n-1):
            if i==0:#append 2nd line in rows
                rows.append([time[i],u[i],v[i],w[i],avg_of_u,avg_of_v,avg_of_w,u_[i],v_[i],w_[i],octant[i],"","Overall Count",total[0],total[1],total[2],total[3],total[4],total[5],total[6],total[7]])
            elif i==1:#append 3nd line in rows
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"User Input","Mod "+ str(mod)])
            elif i==2:#append 4nd line in rows
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"",str((i-2)*mod)+"-"+str((i-1)*mod-1),octant_1[i-2],octant__1[i-2],octant_2[i-2],octant__2[i-2],octant_3[i-2],octant__3[i-2],octant_4[i-2],octant__4[i-2]])
            elif i>2 and i<=2+(n-2)//mod:#append (n-2)//mod lines in rows(represent count of octants in per gap of mod)
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"",str((i-2)*mod)+"-"+str(np.minimum((i-1)*mod-1,n-2)),octant_1[i-2],octant__1[i-2],octant_2[i-2],octant__2[i-2],octant_3[i-2],octant__3[i-2],octant_4[i-2],octant__4[i-2]])
            elif i==3+(n-2)//mod:#append a line alone becouse of one word (Verified)
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"","Verified",np.sum(octant_1),np.sum(octant__1),np.sum(octant_2),np.sum(octant__2),np.sum(octant_3),np.sum(octant__3),np.sum(octant_4),np.sum(octant__4)])
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i]]) #for blank space juse below       
            elif i==4+(n-2)//mod:
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i]])
            elif i==5+(n-2)//mod:#append a line alone becouse of one word (Overall Transition Count)
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"","Overall Transition Count"])
            elif (i-(6+(n-2)//mod))%13==0 and i<((n-2)//mod +2)*14 +4:#append a line alone becouse of one word (To)
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"",line2[i+2-(7+(n-2)//mod)],"To"])
            elif (i-(7+(n-2)//mod))%13==0 and i<((n-2)//mod +2)*14 +4:#append a list which contain "+1","-1","+2","-2","+3","-3","+4","-4" in every 13th line
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"",line2[i+2-(7+(n-2)//mod)],"+1","-1","+2","-2","+3","-3","+4","-4"])
            elif (i-(8+(n-2)//mod))%13>=0 and (i-(8+(n-2)//mod))%13<8 and i<((n-2)//mod +2)*3 +4:
                b = list_[(i-(8+(n-2)//mod))%13]#list_[1,-1,2,-2,3,-3,4,-4]
                print(i-(8+((n-2)//mod)))
                list1 = function1(0,n-2,b)#count b,(1,-1,2,-2,3,-3,4,-4) from 0 to end in octant list and make a list named list1[0]
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],line1[i-(8+(n-2)//mod)],line2[i+2-(7+(n-2)//mod)],list1[0],list1[1],list1[2],list1[3],list1[4],list1[5],list1[6],list1[7]])
            elif (i-(8+(n-2)//mod))%13>=0 and (i-(8+(n-2)//mod))%13<8 and i<((n-2)//mod +2)*14 +4:
                b = list_[(i-(8+(n-2)//mod))%13]#count b,(1,-1,2,-2,3,-3,4,-4) from ((i-(8+(n-2)//mod))//13-1)*mod to min(n-2,((i-(8+(n-2)//mod))//13)*mod-1),b in octant list and make a list named list1[0]
                list1 = function1(((i-(8+(n-2)//mod))//13-1)*mod,min(n-2,((i-(8+(n-2)//mod))//13)*mod),b)
                print(i-(8+((n-2)//mod)))
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],line1[i-(8+((n-2)//mod))],line2[i+2-(7+(n-2)//mod)],list1[0],list1[1],list1[2],list1[3],list1[4],list1[5],list1[6],list1[7]])
            elif i<((n-2)//mod +2)*14 +4:
                print(i-(8+((n-2)//mod)))
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],line1[i-(8+((n-2)//mod))],line2[i+2-(7+(n-2)//mod)]])
            else:
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i]])
    
        for row in rows:
            sheet.append(row)
        book.save('output_octant_transition_identify.xlsx')
    except:
        print("File does not exist")
        exit()
        
mod=5000
octact_transition_count(mod)