from datetime import datetime
start_time = datetime.now()
import numpy as np
def octant_range_names(mod=5000):
        from openpyxl import load_workbook
    #try:
        octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
        wb = load_workbook("octant_input.xlsx")
        sheet = wb["Sheet1"]
        n = sheet.max_row
        #n = row_count
        time =[]#time list
        u = []#U list
        v = []#V list
        w = []#W list
        u_ = []# u - avg_of_u
        v_ = []#v - avg_of_v
        w_ = []#w - avg_of_w
        octant = []#store octant value
        total = [0]*8#count of 1,2,3,4,-1,-2,-3,-4
        list_ = [1,-1,2,-2,3,-3,4,-4]

        for i in range(2, n + 1):
            time.append(sheet.cell(row=i, column=1).value)#append time value in time list
            u.append(sheet.cell(row=i, column=2).value)
            v.append(sheet.cell(row=i, column=3).value)
            w.append(sheet.cell(row=i, column=4).value)

        avg_of_u = np.mean(u)
        avg_of_v = np.mean(v)
        avg_of_w = np.mean(w)#store avg of u,v,w in avg_of_u,avg_of_v,avg_of_w

        for i in range(n-1):
            v_.append(v[i]-avg_of_v)#append V-Vavg in v_
            u_.append(u[i]-avg_of_u)
            w_.append(w[i]-avg_of_w)
        
        for i in range(n-1):#for identfy octant
                if u_[i]>=0 and v_[i]>=0 and w_[i]>0:
                    octant.append(+1)
                elif u_[i]<0 and v_[i]>=0 and w_[i]>0:
                    octant.append(+2)
                elif u_[i]<0 and v_[i]<0 and w_[i]>=0:
                    octant.append(+3)
                elif u_[i]>=0 and v_[i]<0 and w_[i]>=0:
                    octant.append(+4)
                elif u_[i]>=0 and v_[i]>=0 and w_[i]<=0:
                    octant.append(-1)
                elif u_[i]<0 and v_[i]>=0 and w_[i]<=0:
                    octant.append(-2)
                elif u_[i]<0 and v_[i]<0 and w_[i]<0:
                    octant.append(-3)
                else:
                    octant.append(-4)

        for j in range(0,n-1):#for count total no of 1,2,3,4,-1,-2,-3,-4
            if octant[j] ==1:
                total[0] = total[0] +1
            elif octant[j] ==2:
                total[2] = total[2] +1
            elif octant[j] ==3:
                total[4] = total[4] +1
            elif octant[j] ==4:
                total[6] = total[6] +1
            elif octant[j] ==-1:
                total[1] = total[1] +1
            elif octant[j] ==-2:
                total[3] = total[3] +1
            elif octant[j] ==-3:
                total[5] = total[5] +1
            else:
                total[7] = total[7] +1

        rank_1 = []
        
        rank_total = []
        for i in range(8):
            a =1
            for j in range(8):
                if total[i]<total[j]:
                    a = a+1
            if a==1:
                rank_1.append(list_[i])
            rank_total.append(a)
                
        octant_2d = [[0]*((n-2)//mod+1),[0]*((n-2)//mod+1),[0]*((n-2)//mod+1),[0]*((n-2)//mod+1),[0]*((n-2)//mod+1),[0]*((n-2)//mod+1),[0]*((n-2)//mod+1),[0]*((n-2)//mod+1)]#store count of 1 in every mod value
            
        for i in range(0,(n-2)//mod):
            for j in range(i*mod,(i+1)*mod):
                if octant[j] ==1:#count of 1 in i*mod_1 to (i+1)*mod
                    octant_2d[0][i] = octant_2d[0][i] +1
                elif octant[j] ==2:
                    octant_2d[2][i] = octant_2d[2][i] +1
                elif octant[j] ==3:
                    octant_2d[4][i] = octant_2d[4][i] +1
                elif octant[j] ==4:
                    octant_2d[6][i] = octant_2d[6][i] +1
                elif octant[j] ==-1:
                    octant_2d[1][i] = octant_2d[1][i] +1
                elif octant[j] ==-2:
                    octant_2d[3][i] = octant_2d[3][i] +1
                elif octant[j] ==-3:
                    octant_2d[5][i] = octant_2d[5][i] +1
                else:
                    octant_2d[7][i] = octant_2d[7][i] +1

        for i in range(((n-2)//mod)*mod,n-1):
            if octant[i] ==1:#count 1 in rest element 
                octant_2d[0][(n-2)//mod] = octant_2d[0][(n-2)//mod] +1
            elif octant[i] ==2:
                octant_2d[2][(n-2)//mod] = octant_2d[2][(n-2)//mod] +1
            elif octant[i] ==3:
                octant_2d[4][(n-2)//mod] = octant_2d[4][(n-2)//mod] +1
            elif octant[i] ==4:
                octant_2d[6][(n-2)//mod] = octant_2d[6][(n-2)//mod] +1
            elif octant[i] ==-1:
                octant_2d[1][(n-2)//mod] = octant_2d[1][(n-2)//mod] +1
            elif octant[i] ==-2:
                octant_2d[3][(n-2)//mod] = octant_2d[3][(n-2)//mod] +1
            elif octant[i] ==-3:
                octant_2d[5][(n-2)//mod] = octant_2d[5][(n-2)//mod] +1
            else:
                octant_2d[7][(n-2)//mod] = octant_2d[7][(n-2)//mod] +1  

        rank_2d = []
        for i in range(((n-2)//mod)+1):
            rank_mod = []
            for j in range(8):
                a = 1
                for k in range(8):
                    if octant_2d[j][i]<octant_2d[k][i]:
                        a = a+1
                if a==1:
                    rank_1.append(list_[j])
                rank_mod.append(a)
            rank_2d.append(rank_mod)
        
        from openpyxl import Workbook
        book = Workbook()
        sheet = book.active

        rows = [
            ["","","","","","","","","","","","","","","","","","","","","","+1","-1","+2","-2","+3","-3","+4","-4"],["Time","U","V","W","U Avg","V Avg","W Avg","U'=U - U avg","V'=V - V avg","W'=W - w avg","Ocatant","","","+1","-1","+2","-2","+3","-3","+4","-4","Rank 1","Rank 2","Rank 3","Rank 4","Rank 5","Rank 6","Rank 7","Rank 8","Rank1 Octant ID","Rank1 Octant Name"]
        ]#made 2d list

        for i in range(n-1):
            if i==0:#append 2nd line in rows
                rows.append([time[i],u[i],v[i],w[i],avg_of_u,avg_of_v,avg_of_w,u_[i],v_[i],w_[i],octant[i],"","Overall Count",total[0],total[1],total[2],total[3],total[4],total[5],total[6],total[7],rank_total[0],rank_total[1],rank_total[2],rank_total[3],rank_total[4],rank_total[5],rank_total[6],rank_total[7],rank_1[0],octant_name_id_mapping[str(rank_1[0])]])
            elif i==1:#append 3nd line in rows
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"User Input","Mod "+ str(mod)])
            elif i>=2 and i<=2+(n-2)//mod:#append (n-2)//mod lines in rows(represent count of octants in per gap of mod)
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"",str((i-2)*mod)+"-"+str(np.minimum((i-1)*mod-1,n-2)),octant_2d[0][i-2],octant_2d[1][i-2],octant_2d[2][i-2],octant_2d[3][i-2],octant_2d[4][i-2],octant_2d[5][i-2],octant_2d[6][i-2],octant_2d[7][i-2],rank_2d[i-2][0],rank_2d[i-2][1],rank_2d[i-2][2],rank_2d[i-2][3],rank_2d[i-2][4],rank_2d[i-2][5],rank_2d[i-2][6],rank_2d[i-2][7],rank_1[i-1],octant_name_id_mapping[str(rank_1[i-1])]])     
            elif i ==6+(n-2)//mod:
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"","","Octant ID","Octant Name","Count of Rank 1 Mod Values"])
            elif i >=7+(n-2)//mod and i<=14+(n-2)//mod:
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i],"","",list_[i-(7+(n-2)//mod)],octant_name_id_mapping[str(list_[i-(7+(n-2)//mod)])]])
            else:
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i]])
    
        for row in rows:
            sheet.append(row)
        book.save('output_octant_ranking_excel.xlsx')
    #except:
        #print("File does not exist")
        #exit()
mod=5000 
octant_range_names(mod)
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
