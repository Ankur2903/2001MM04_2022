import numpy as np
from openpyxl import load_workbook
def octant_longest_subsequence_count():
    try:
        wb = load_workbook("input_octant_longest_subsequence.xlsx")
        sheet = wb["Sheet1"]
        n = sheet.max_row
        #n = row_count
        time =[]#time list
        u = []#U list
        v = []#V list
        w = []#W list
        u_ = []# u - avg_of_u
        v_ = []#v - avg_of_v
        w_ = []#w - avg_of_w
        octant = []#store octant value
        octants = ["+1","-1","+2","-2","3","-3","+4","-4"]

        for i in range(2, n + 1):
            time.append(sheet.cell(row=i, column=1).value)#append time value in time list
            u.append(sheet.cell(row=i, column=2).value)
            v.append(sheet.cell(row=i, column=3).value)
            w.append(sheet.cell(row=i, column=4).value)

        avg_of_u = np.mean(u)
        avg_of_v = np.mean(v)
        avg_of_w = np.mean(w)#store avg of u,v,w in avg_of_u,avg_of_v,avg_of_w

        for i in range(n-1):
            v_.append(v[i]-avg_of_v)#append V-Vavg in v_
            u_.append(u[i]-avg_of_u)
            w_.append(w[i]-avg_of_w)

        for i in range(n-1):#for identfy octant
            if u_[i]>=0 and v_[i]>=0 and w_[i]>0:
                octant.append(+1)
            elif u_[i]<0 and v_[i]>=0 and w_[i]>0:
                octant.append(+2)
            elif u_[i]<0 and v_[i]<0 and w_[i]>=0:
                octant.append(+3)
            elif u_[i]>=0 and v_[i]<0 and w_[i]>=0:
                octant.append(+4)
            elif u_[i]>=0 and v_[i]>=0 and w_[i]<=0:
                octant.append(-1)
            elif u_[i]<0 and v_[i]>=0 and w_[i]<=0:
                octant.append(-2)
            elif u_[i]<0 and v_[i]<0 and w_[i]<0:
                octant.append(-3)
            else:
                octant.append(-4)
        
        from openpyxl import Workbook
        book = Workbook()
        sheet = book.active

        rows = [
            ["Time","U","V","W","U Avg","V Avg","W Avg","U'=U - U avg","V'=V - V avg","W'=W - w avg","Ocatant","","Count","Longest Subsquence Length","Count"]
        ]#made 2d list

        for i in range(n-1):
            if i==0:#append 2nd line in rows
                rows.append([time[i],u[i],v[i],w[i],avg_of_u,avg_of_v,avg_of_w,u_[i],v_[i],w_[i],octant[i]])
            elif i<9:
                rows.append([time[i],u[i],v[i],w[i],avg_of_u,avg_of_v,avg_of_w,u_[i],v_[i],w_[i],octant[i],"",octants[i-1]])
            else:
                rows.append([time[i],u[i],v[i],w[i]," "," "," ",u_[i],v_[i],w_[i],octant[i]])
    
        for row in rows:
            sheet.append(row)
        book.save('output_octant_longest_subsequence.xlsx')
    except:
        print("File does not exist")
        exit()
octant_longest_subsequence_count()